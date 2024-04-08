#!/usr/bin/env python3
import argparse
from openpyxl import load_workbook
import io
import os
import requests
from collections import OrderedDict
import subprocess
import shutil
import re
import json

parser = argparse.ArgumentParser(description='Run with no option to generate files to /tmp or point to the mscp directory')
parser.add_argument("-r", "--repo", default="/tmp/", help="Directory mscp is cloned to", type=str)
results = parser.parse_args()

if results.repo != "/tmp/":
    if results.repo[-1] != "/":
        path = results.repo + "/build"
    else:
        path = results.repo + "/build"
else:
    path = results.repo

latest_url = 'https://api.github.com/repos/securecontrolsframework/securecontrolsframework/releases/latest'
latest_r = requests.get(latest_url, allow_redirects=True)
latest_response_data = json.loads(latest_r.content.decode('utf-8'))
download_url = ""
url = 'https://api.github.com/repos/securecontrolsframework/securecontrolsframework/contents'
r = requests.get(url, allow_redirects=True)
response_data = json.loads(r.content.decode('utf-8'))

for item in response_data:
    for k,v in item.items():
        if k == "download_url":    
            try:
                if "xlsx" in v:
                    download_url = v
            except:
                continue
if download_url == "":
    print("No xlsx file found")
    exit()
r = requests.get(download_url,allow_redirects=True)
open(path + '/SCF_current.xlsx', 'wb').write(r.content)
workbook = load_workbook(filename=path + "/SCF_current.xlsx")
sheet = workbook['{}'.format(latest_response_data['name'])]
frameworklist = [""]
for cell in sheet[1]:    
    if cell.fill.start_color.index == 5 or cell.fill.start_color.index == 9 or cell.fill.start_color.index == 4 or cell.fill.start_color.index == 3:
        if "Risk" not in str(cell.value) and "800-53" not in str(cell.value) and "800-171" not in str(cell.value) and "CIS" not in str(cell.value) and "SP-CMM" not in str(cell.value) and "CMMC" not in str(cell.value):
            # print()
            frameworklist.append(str(cell.value).replace('\n'," "))

frameworklist.sort()
for framework in frameworklist[1:]:
    print("{}. {}".format(frameworklist.index(framework),framework))
        
print()
framework_number = input("Enter Number for Framework for Mapping: ")
framework = frameworklist[int(framework_number)]

mapping = {}
keys = []
values = []

print("Generating mapping for {}".format(framework))
row_array = []
nist_column = int()
for column_cell in sheet.iter_cols(1, sheet.max_column):  # iterate column cell

    if str(column_cell[0].value).replace("\n"," ") == framework:    # check for your column
        keys.append(framework.replace(" ","_"))
        for data in column_cell[1:]:    # iterate your column
            if data.value == None:
                continue
            else:
                keys.append("\"" + str(data.value).replace("\n",", ") + "\"")
                row_array.append(data.row)
        break   

for column_cell in sheet.iter_cols(1, sheet.max_column):
    if str(column_cell[0].value).replace("\n"," ") == "NIST 800-53 rev5":
        values.append("800-53r5")
        for data in column_cell[1:]:
            if data.row in row_array:
                if data.value == None:
                    values.append("N/A")
                else:
                    values.append("\"" + str(data.value).replace("\n",", ") + "\"")

fullcsv = str()
missingcontrols = str()

counter = 0
for key in keys:
    fullcsv = fullcsv + key + "," + values[counter] + "\n"
    if values[counter] == "N/A":
        missingcontrols = missingcontrols + key + "\n"
    counter += 1

framework_filename = framework.replace(" ","_").replace("(","_").replace(")","_")
path_to_framework_mapping = path + "/" + framework_filename + "-mapping.csv"
missing_controls = path + "/" + framework_filename + "-missingcontrols.txt"
with open(path_to_framework_mapping,'w') as rite:
        rite.write(fullcsv)
rite.close()

with open(missing_controls,'w') as rite:
        rite.write(missingcontrols)
rite.close()

print("{} and {} written to {}".format(framework_filename + "-mapping.csv", framework_filename + "-missingcontrols.txt", path))

if results.repo != "/tmp/":
    
    script_path = path[:-6]
    full_path_mapping = os.path.abspath(path_to_framework_mapping)
    subprocess.call(script_path + "/scripts/generate_mapping.py " + full_path_mapping , shell=True)
    ogpath = script_path + "/build/" + framework.replace(" ","_") + "/rules/"
    rules_dir = os.listdir(ogpath)
    for section in rules_dir:
        original = ogpath + "/" + section
        target = script_path + "/custom/rules/" + section
    
        if os.path.isdir(target):
            existing_custom_rules = os.listdir(target)
            section = str(target).split("/")[-1]
            matched_rules = []
            for existing_custom_rule in existing_custom_rules:
                if existing_custom_rule in os.listdir(ogpath + section):
                    matched_rules.append(ogpath + section + "/" + existing_custom_rule)

            for new_matched_custom in matched_rules:
            
                with open(new_matched_custom) as reader:
                    new_custom_yam = reader.read()
            
                regex = r"(?s)(?<=custom\:).*?(?=tags\:)"
                ref_match = re.search(regex,new_custom_yam)
                regex = r"(?s)(?<=tags\:).*?($)"
                tag_match = re.search(regex,new_custom_yam)
                
                with open(target + "/" + str(new_matched_custom).split("/")[-1]) as reader:
                    existing_custom_yam = reader.read()
                
                existing_custom_yam = existing_custom_yam.replace("custom:","custom:{}".format(ref_match.group(0).rstrip()))
                existing_custom_yam = existing_custom_yam.replace("tags:","tags:{}".format(tag_match.group(0).rstrip()))
                
                with open(target + "/" + str(new_matched_custom).split("/")[-1], 'w') as rite:
                    rite.write(existing_custom_yam)  
    
            for rule in os.listdir(ogpath + section):
                if not os.path.exists(target + "/" + rule):
                    shutil.copyfile(ogpath + section + "/" + rule, target + "/" + rule)

        elif not os.path.isdir(target):
            # fullpath = os.path.abspath(target)
            shutil.move(original, target)        
        
    custom_baseline_file = script_path + "/build/" + framework.replace(" ","_") + "/baseline/" + framework.lower().replace(" ","_") + ".yaml"
    custom_baseline_file = custom_baseline_file.replace(" ","\ ").replace("(","\(").replace(")","\)")
    full_path_baseline = os.path.abspath(custom_baseline_file)
    print(script_path + "/scripts/generate_guidance.py -p -x -s " + full_path_baseline)
    subprocess.call(script_path + "/scripts/generate_guidance.py -p -x -s " + full_path_baseline , shell=True)
